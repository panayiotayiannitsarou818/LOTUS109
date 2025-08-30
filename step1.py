#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Βήμα 1 (IMMUTABLE): Κατανομή παιδιών εκπαιδευτικών με φιλίες
Δημιουργεί ΒΗΜΑ1_ΣΕΝΑΡΙΟ_X στήλες που είναι ΜΟΝΟ ΓΙΑ ΑΝΑΓΝΩΣΗ

ΚΑΝΟΝΑΣ IMMUTABILITY: Όταν δημιουργηθούν οι στήλες ΒΗΜΑ1_ΣΕΝΑΡΙΟ_X,
είναι ΟΡΙΣΤΙΚΕΣ και δεν αλλάζουν ποτέ στα επόμενα βήματα.
"""

from dataclasses import dataclass, field
from typing import Dict, List, Set, Tuple, Optional, FrozenSet
import pandas as pd
import numpy as np
import itertools
import math

# Σταθερά: η στήλη του σεναρίου στο Excel είναι ΠΑΝΤΑ η 'K'
SCENARIO_COL_LETTER = 'K'
import re
import ast
from pathlib import Path


@dataclass(frozen=True)
class Step1Scenario:
    """Immutable σενάριο βήματος 1"""
    id: int
    column_name: str  # "ΒΗΜΑ1_ΣΕΝΑΡΙΟ_1"
    assignments: Dict[str, str]  # name -> class
    description: str
    broken_friendships: int
    metadata: Dict[str, any] = field(default_factory=dict)
    
    def get_assignment(self, student_name: str) -> Optional[str]:
        """Read-only πρόσβαση σε ανάθεση"""
        return self.assignments.get(student_name)
    
    def get_students_in_class(self, class_name: str) -> List[str]:
        """Επιστρέφει λίστα μαθητών σε τμήμα"""
        return [name for name, cls in self.assignments.items() if cls == class_name]


@dataclass(frozen=True)
class Step1Results:
    """Immutable αποτελέσματα βήματος 1"""
    scenarios: Tuple[Step1Scenario, ...]
    friendships: FrozenSet[Tuple[str, str]]
    teacher_kids: Tuple[str, ...]
    num_classes: int
    creation_timestamp: str
    
    def get_scenario(self, scenario_id: int) -> Optional[Step1Scenario]:
        """Επιστρέφει σενάριο με βάση ID"""
        for scenario in self.scenarios:
            if scenario.id == scenario_id:
                return scenario
        return None
    
    def get_scenario_by_column(self, column_name: str) -> Optional[Step1Scenario]:
        """Επιστρέφει σενάριο με βάση όνομα στήλης"""
        for scenario in self.scenarios:
            if scenario.column_name == column_name:
                return scenario
        return None
    
    def validate_immutability(self, df: pd.DataFrame) -> bool:
        """Ελέγχει ότι οι στήλες ΒΗΜΑ1_ΣΕΝΑΡΙΟ_X δεν έχουν αλλάξει"""
        for scenario in self.scenarios:
            col_name = scenario.column_name
            if col_name not in df.columns:
                raise ValueError(f"Λείπει στήλη {col_name} - παραβίαση immutability")
            
            # Έλεγχος ότι οι αναθέσεις είναι οι αναμενόμενες
            for student_name, expected_class in scenario.assignments.items():
                student_row = df[df["ΟΝΟΜΑ"] == student_name]
                if student_row.empty:
                    continue
                
                actual_class = student_row[col_name].iloc[0]
                if pd.notna(actual_class) and str(actual_class).strip() != expected_class:
                    raise ValueError(
                        f"ΠΑΡΑΒΙΑΣΗ IMMUTABILITY: {student_name} σε {col_name} "
                        f"αναμενόταν '{expected_class}', βρέθηκε '{actual_class}'"
                    )
        return True


class Step1ImmutableProcessor:
    """Επεξεργαστής που εξασφαλίζει immutability του Βήματος 1"""
    
    def __init__(self):
        self._results: Optional[Step1Results] = None
        self._is_locked: bool = False
    
    def create_scenarios(self, df: pd.DataFrame, num_classes: Optional[int] = None) -> Step1Results:
        """Δημιουργία immutable σεναρίων"""
        if self._is_locked:
            raise RuntimeError("Step1 είναι ήδη κλειδωμένο - δεν επιτρέπονται αλλαγές")
        
        # Φόρτωση και normalization δεδομένων
        df_norm = self._normalize_dataframe(df)
        
        # Αυτόματος υπολογισμός τμημάτων
        if num_classes is None:
            num_classes = max(2, math.ceil(len(df_norm) / 25))
        
        print(f"Βήμα 1 - Δημιουργία immutable σεναρίων για {num_classes} τμήματα")
        
        # Εντοπισμός παιδιών εκπαιδευτικών
        teacher_kids = self._get_teacher_kids(df_norm)
        if not teacher_kids:
            print("Δεν υπάρχουν παιδιά εκπαιδευτικών - κενά αποτελέσματα")
            return Step1Results(
                scenarios=tuple(),
                friendships=frozenset(),
                teacher_kids=tuple(),
                num_classes=num_classes,
                creation_timestamp=pd.Timestamp.now().isoformat()
            )
        
        print(f"Εντοπίστηκαν {len(teacher_kids)} παιδιά εκπαιδευτικών")
        
        # Εξαγωγή φιλιών
        friendships = self._extract_friendships(df_norm, teacher_kids)
        
        # Δημιουργία σεναρίων
        scenarios = self._generate_scenarios(teacher_kids, num_classes, friendships)
        
        # Δημιουργία immutable αποτελεσμάτων
        self._results = Step1Results(
            scenarios=tuple(scenarios),
            friendships=friendships,
            teacher_kids=tuple(teacher_kids),
            num_classes=num_classes,
            creation_timestamp=pd.Timestamp.now().isoformat()
        )
        
        print(f"Δημιουργήθηκαν {len(scenarios)} immutable σενάρια")
        return self._results
    
    def apply_to_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Εφαρμόζει τα σενάρια στο DataFrame ΚΑΙ το κλειδώνει"""
        if not self._results:
            raise RuntimeError("Δεν έχουν δημιουργηθεί σενάρια ακόμη")
        
        result_df = df.copy()
        
        # Προσθήκη στηλών ΒΗΜΑ1_ΣΕΝΑΡΙΟ_X
        for scenario in self._results.scenarios:
            col_name = scenario.column_name
            result_df[col_name] = ""  # Αρχικοποίηση με κενές τιμές
            
            # Συμπλήρωση μόνο για παιδιά εκπαιδευτικών
            for student_name, class_assigned in scenario.assignments.items():
                mask = result_df["ΟΝΟΜΑ"] == student_name
                if mask.any():
                    result_df.loc[mask, col_name] = class_assigned
        
        # ΚΛΕΙΔΩΜΑ - μετά από αυτό δεν επιτρέπονται αλλαγές
        self._is_locked = True
        print(f"ΚΛΕΙΔΩΜΑ: Οι στήλες {[s.column_name for s in self._results.scenarios]} είναι πλέον IMMUTABLE")
        
        return result_df
    
    def get_results(self) -> Optional[Step1Results]:
        """Read-only πρόσβαση στα αποτελέσματα"""
        return self._results
    
    def is_locked(self) -> bool:
        """Έλεγχος αν το Step1 είναι κλειδωμένο"""
        return self._is_locked
    
    def validate_external_dataframe(self, df: pd.DataFrame) -> bool:
        """Επικυρώνει ότι εξωτερικό DataFrame τηρεί την immutability"""
        if not self._results:
            raise RuntimeError("Δεν υπάρχουν αποτελέσματα προς επικύρωση")
        
        return self._results.validate_immutability(df)
    
    # === ΒΟΗΘΗΤΙΚΕΣ ΜΕΘΟΔΟΙ (από step1_friendships.py) ===
    
    def _normalize_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Κανονικοποίηση DataFrame"""
        result = df.copy()
        
        # Κανονικοποίηση ονομάτων στηλών
        rename = {}
        for c in result.columns:
            cc = str(c).strip()
            if cc.lower() in ["ονομα", "name", "μαθητης", "μαθητρια"]:
                rename[c] = "ΟΝΟΜΑ"
            elif cc.lower().startswith("φυλο") or cc.lower() == "gender":
                rename[c] = "ΦΥΛΟ"
            elif "γνωση" in cc.lower():
                rename[c] = "ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ"
            elif "εκπ" in cc.lower():
                rename[c] = "ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ"
        
        result.rename(columns=rename, inplace=True)
        
        # Έλεγχος απαιτούμενων στηλών
        required_cols = ["ΟΝΟΜΑ", "ΦΥΛΟ", "ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ", "ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ"]
        for col in required_cols:
            if col not in result.columns:
                raise ValueError(f"Λείπει απαιτούμενη στήλη: {col}")
        
        # Κανονικοποίηση τιμών
        result["ΟΝΟΜΑ"] = result["ΟΝΟΜΑ"].astype(str).str.strip()
        result["ΦΥΛΟ"] = result["ΦΥΛΟ"].astype(str).str.strip().str.upper().map({
            "Α": "Α", "Κ": "Κ", "AGORI": "Α", "KORITSI": "Κ"
        }).fillna("")
        
        for c in ["ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ", "ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ"]:
            result[c] = result[c].map(self._norm_yesno)
        
        return result
    
    def _norm_yesno(self, val) -> str:
        """Κανονικοποίηση Ν/Ο τιμών"""
        s = str(val).strip().upper()
        return "Ν" if s in {"Ν", "ΝΑΙ", "YES", "TRUE", "1", "Y"} else "Ο"
    
    def _get_teacher_kids(self, df: pd.DataFrame) -> List[str]:
        """Εντοπισμός παιδιών εκπαιδευτικών"""
        teacher_kids_df = df[df["ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ"] == "Ν"]
        return teacher_kids_df["ΟΝΟΜΑ"].astype(str).str.strip().tolist()
    
    def _find_friendship_columns(self, df: pd.DataFrame) -> List[str]:
        """Εντοπισμός στηλών φιλιών (στήλες με ονόματα μαθητών)"""
        student_names = set(df["ΟΝΟΜΑ"].astype(str).str.strip())
        friendship_cols = []
        
        for col in df.columns:
            col_name = str(col).strip()
            if col_name in student_names and col_name != "ΟΝΟΜΑ":
                friendship_cols.append(col)
        
        return friendship_cols
    
    def _extract_friendships(self, df: pd.DataFrame, teacher_kids: List[str]) -> FrozenSet[Tuple[str, str]]:
        """Εξαγωγή αμοιβαίων φιλιών μεταξύ παιδιών εκπαιδευτικών"""
        teacher_kids_set = set(teacher_kids)
        student_friends = {}
        
        # ΜΕΘΟΔΟΣ 1: Matrix-style (στήλες με ονόματα)
        friendship_cols = self._find_friendship_columns(df)
        if friendship_cols:
            print(f"Εντοπίστηκαν {len(friendship_cols)} στήλες φιλιών (matrix-style)")
            
            for col in friendship_cols:
                friend_name = str(col).strip()
                if friend_name in teacher_kids_set:
                    # Βρες ποια παιδιά εκπαιδευτικών έχουν γράψει το friend_name ως φίλο
                    friend_mask = (df["ΟΝΟΜΑ"].isin(teacher_kids)) & (df[col].map(self._norm_yesno) == "Ν")
                    students_who_wrote_this_friend = df[friend_mask]["ΟΝΟΜΑ"].tolist()
                    
                    for student in students_who_wrote_this_friend:
                        if student != friend_name:  # Όχι φιλία με τον εαυτό του
                            if student not in student_friends:
                                student_friends[student] = set()
                            student_friends[student].add(friend_name)
        
        # ΜΕΘΟΔΟΣ 2: Single-column ΦΙΛΟΙ (fallback)
        elif "ΦΙΛΟΙ" in df.columns:
            print("Χρήση στήλης ΦΙΛΟΙ (single-column)")
            
            for _, row in df.iterrows():
                student_name = str(row["ΟΝΟΜΑ"]).strip()
                if student_name in teacher_kids_set:
                    friends_str = str(row["ΦΙΛΟΙ"]).strip()
                    if friends_str and friends_str.lower() not in ["", "nan", "none"]:
                        # Split με διάφορα separators
                        friends_list = []
                        for sep in [",", ";", "|"]:
                            if sep in friends_str:
                                friends_list = [f.strip() for f in friends_str.split(sep)]
                                break
                        else:
                            friends_list = [friends_str.strip()]  # Single friend
                        
                        # Φιλτράρισμα μόνο παιδιών εκπαιδευτικών
                        valid_friends = [f for f in friends_list if f in teacher_kids_set and f != student_name]
                        
                        if valid_friends:
                            student_friends[student_name] = set(valid_friends)
        
        else:
            print("Δεν βρέθηκαν στήλες φιλιών")
        
        # Έλεγχος αμοιβαιότητας: A→B ΚΑΙ B→A
        friendships = set()
        for student_a in student_friends:
            friends_of_a = student_friends[student_a]
            for student_b in friends_of_a:
                if student_b in student_friends and student_a in student_friends[student_b]:
                    # Αμοιβαία φιλία βρέθηκε - προσθήκη με sorted order
                    pair = tuple(sorted([student_a, student_b]))
                    friendships.add(pair)
        
        print(f"Βρέθηκαν {len(friendships)} αμοιβαίες φιλίες μεταξύ παιδιών εκπαιδευτικών")
        return frozenset(friendships)
    
    def _count_broken_friendships(self, teacher_kids: List[str], assign_map: Dict[str, str], 
                               friendships: FrozenSet[Tuple[str, str]]) -> int:
        """Μέτρηση σπασμένων φιλιών σε ένα σενάριο κατανομής"""
        broken = 0
        for friend1, friend2 in friendships:
            if assign_map.get(friend1) != assign_map.get(friend2):
                broken += 1
        return broken
    
    def _canonical_key(self, names: List[str], assign_map: Dict[str, str], class_labels_list: List[str]) -> Tuple:
        """Canonical key για αποφυγή duplicates"""
        buckets = []
        for c in class_labels_list:
            members = tuple(sorted([n for n in names if assign_map.get(n) == c]))
            buckets.append(members)
        return tuple(sorted(buckets))
    
    def _generate_scenarios(self, teacher_kids: List[str], num_classes: int, 
                          friendships: FrozenSet[Tuple[str, str]]) -> List[Step1Scenario]:
        """Δημιουργία σεναρίων με immutable structure"""
        class_labels_list = [f"Α{i+1}" for i in range(num_classes)]
        scenarios = []
        
        if len(teacher_kids) <= num_classes:
            # ΚΑΝΟΝΑΣ 1: Σειριακή κατανομή
            print(f"Εφαρμογή Κανόνα 1 (≤1 ανά τμήμα)")
            assignments = {}
            
            for i, name in enumerate(teacher_kids):
                assignments[name] = class_labels_list[i % num_classes]
            
            scenario = Step1Scenario(
                id=1,
                column_name="ΒΗΜΑ1_ΣΕΝΑΡΙΟ_1",
                assignments=assignments,
                description="Κανόνας 1: Σειριακή κατανομή ≤1/τμήμα",
                broken_friendships=0
            )
            scenarios.append(scenario)
        else:
            # ΚΑΝΟΝΑΣ 2: Εξαντλητική παραγωγή
            print(f"Εφαρμογή Κανόνα 2 (εξαντλητική με φιλίες)")
            valid_assignments = self._exhaustive_generation(teacher_kids, num_classes, friendships)
            
            for i, (assignments_dict, broken_count) in enumerate(valid_assignments[:5], 1):
                scenario = Step1Scenario(
                    id=i,
                    column_name=f"ΒΗΜΑ1_ΣΕΝΑΡΙΟ_{i}",
                    assignments=assignments_dict,
                    description="Κανόνας 2: Ισόρροπη κατανομή",
                    broken_friendships=broken_count
                )
                scenarios.append(scenario)
        
        return scenarios
    
    def _exhaustive_generation(self, teacher_kids: List[str], num_classes: int, 
                             friendships: FrozenSet[Tuple[str, str]]) -> List[Tuple[Dict[str, str], int]]:
        """Εξαντλητική παραγωγή σεναρίων"""
        class_labels_list = [f"Α{i+1}" for i in range(num_classes)]
        valid_scenarios = []
        seen_canonical = set()
        
        print(f"Παραγωγή σεναρίων για {len(teacher_kids)} παιδιά σε {num_classes} τμήματα...")
        
        # Εξαντλητική παραγωγή
        total_combinations = num_classes ** len(teacher_kids)
        print(f"Συνολικές περιπτώσεις: {total_combinations:,}")
        
        for assignment in itertools.product(class_labels_list, repeat=len(teacher_kids)):
            assign_map = {teacher_kids[i]: assignment[i] for i in range(len(teacher_kids))}
            
            # ΕΛΕΓΧΟΣ 1: Ισοκατανομή ≤1
            class_counts = {c: 0 for c in class_labels_list}
            for name in teacher_kids:
                class_counts[assign_map[name]] += 1
            
            counts_list = list(class_counts.values())
            if max(counts_list) - min(counts_list) > 1:
                continue  # Απόρριψη ανισοκατανομής >1
            
            # ΕΛΕΓΧΟΣ 2: Όχι όλα στο ίδιο τμήμα
            unique_classes = set(assign_map.values())
            if len(unique_classes) == 1:
                continue  # Απόρριψη
            
            # ΕΛΕΓΧΟΣ 3: Canonical uniqueness
            canon_key = self._canonical_key(teacher_kids, assign_map, class_labels_list)
            if canon_key in seen_canonical:
                continue
            seen_canonical.add(canon_key)
            
            # Υπολογισμός σπασμένων φιλιών
            broken_friendships = self._count_broken_friendships(teacher_kids, assign_map, friendships)
            
            valid_scenarios.append((assign_map, broken_friendships))
        
        print(f"Έγκυρα σενάρια: {len(valid_scenarios)}")
        
        # Φιλτράρισμα αν >5
        if len(valid_scenarios) > 5:
            print("Εφαρμογή φιλτραρίσματος...")
            
            # Προτεραιότητα σε σενάρια με λιγότερα σπασμένα φιλιά
            min_broken = min(s[1] for s in valid_scenarios)
            if min_broken == 0:
                scenarios_without_breaks = [s for s in valid_scenarios if s[1] == 0]
                print(f"Βρέθηκαν {len(scenarios_without_breaks)} σενάρια χωρίς σπασμένες φιλίες")
                valid_scenarios = scenarios_without_breaks
            else:
                print(f"Όλα σπάζουν φιλίες (min: {min_broken}) - ταξινόμηση")
                valid_scenarios.sort(key=lambda x: x[1])
            
            # Τελική επιλογή 5 σεναρίων
            if len(valid_scenarios) > 5:
                valid_scenarios = valid_scenarios[:5]
        
        print(f"Τελική επιλογή: {len(valid_scenarios)} σενάρια")
        return valid_scenarios


# === UTILITY FUNCTIONS ===

def create_immutable_step1(df: pd.DataFrame, num_classes: Optional[int] = None) -> Tuple[pd.DataFrame, Step1Results]:
    """
    Δημιουργεί immutable αποτελέσματα βήματος 1.
    
    Args:
        df: Αρχικό DataFrame με δεδομένα μαθητών
        num_classes: Αριθμός τμημάτων (αν None, αυτόματος υπολογισμός)
    
    Returns:
        (DataFrame με στήλες ΒΗΜΑ1_ΣΕΝΑΡΙΟ_X, Step1Results object)
    """
    processor = Step1ImmutableProcessor()
    results = processor.create_scenarios(df, num_classes)
    updated_df = processor.apply_to_dataframe(df)
    
    return updated_df, results


def validate_step1_immutability(df: pd.DataFrame, results: Step1Results) -> bool:
    """Επικυρώνει ότι το DataFrame τηρεί την immutability του Step1"""
    try:
        return results.validate_immutability(df)
    except ValueError as e:
        print(f"ΣΦΑΛΜΑ IMMUTABILITY: {e}")
        return False


def save_immutable_step1_results(df_with_step1: pd.DataFrame, results: Step1Results, 
                                output_file: str = "ΒΗΜΑ1_IMMUTABLE.xlsx"):
    """Αποθήκευση immutable αποτελεσμάτων"""
    output_path = Path(output_file)
    
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Κύριο sheet με όλα τα δεδομένα + στήλες ΒΗΜΑ1_ΣΕΝΑΡΙΟ_X
        df_with_step1.to_excel(writer, index=False, sheet_name="ΒΗΜΑ1_IMMUTABLE")
        
        # Summary sheet
        summary_data = []
        for scenario in results.scenarios:
            summary_data.append({
                "ID": scenario.id,
                "Στήλη": scenario.column_name,
                "Περιγραφή": scenario.description,
                "Σπασμένες_Φιλίες": scenario.broken_friendships,
                "Αναθέσεις": len(scenario.assignments),
                "Κλείδωμα": "IMMUTABLE"
            })
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, index=False, sheet_name="ΣΕΝΑΡΙΑ_SUMMARY")
    
    print(f"Αποθηκεύτηκε: {output_path}")


# === MAIN EXECUTION ===

def main():
    """Κύρια συνάρτηση για δοκιμή"""
    # Παράδειγμα χρήσης
    try:
        # Υπόθεση: έχουμε ένα sample DataFrame
        sample_data = {
            "ΟΝΟΜΑ": ["Μαρία", "Γιάννης", "Ελένη", "Νίκος"],
            "ΦΥΛΟ": ["Κ", "Α", "Κ", "Α"],
            "ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ": ["Ν", "Ο", "Ν", "Ν"],
            "ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ": ["Ν", "Ν", "Ο", "Ο"]
        }
        df = pd.DataFrame(sample_data)
        
        print("=== ΔΗΜΙΟΥΡΓΙΑ IMMUTABLE STEP1 ===")
        df_with_step1, step1_results = create_immutable_step1(df, num_classes=2)
        
        print("\n=== ΕΠΙΚΥΡΩΣΗ IMMUTABILITY ===")
        is_valid = validate_step1_immutability(df_with_step1, step1_results)
        print(f"Immutability check: {'PASS' if is_valid else 'FAIL'}")
        
        print("\n=== ΑΠΟΘΗΚΕΥΣΗ ===")
        save_immutable_step1_results(df_with_step1, step1_results)
        
        print("\n=== ΣΤΟΙΧΕΙΑ ΣΕΝΑΡΙΩΝ ===")
        for scenario in step1_results.scenarios:
            print(f"{scenario.column_name}: {scenario.assignments}")
        
    except Exception as e:
        print(f"Σφάλμα: {e}")
        raise


if __name__ == "__main__":
    main()


# === WRAPPER FUNCTION ΓΙΑ ΤΑ PAGES ===

def run_step1(df: pd.DataFrame) -> Dict[str, any]:
    """
    Wrapper function που καλούν τα pages.
    Μετατρέπει το immutable API σε format που περιμένουν τα pages.
    """
    try:
        # Κλήση του immutable API
        df_result, step1_results = create_immutable_step1(df)
        
        # Μετατροπή σε format που περιμένουν τα pages
        return {
            "df": df_result,
            "scenarios": {
                "teachers_distributed": len(step1_results.teacher_kids),
                "scenarios_count": len(step1_results.scenarios),
                "broken_friendships": sum(s.broken_friendships for s in step1_results.scenarios)
            },
            "meta": {
                "step": 1,
                "description": "Παιδιά εκπαιδευτικών",
                "teacher_kids_count": len(step1_results.teacher_kids),
                "num_classes": step1_results.num_classes,
                "friendships_found": len(step1_results.friendships)
            }
        }
        
    except Exception as e:
        print(f"Σφάλμα στο run_step1: {e}")
        return {
            "df": df.copy(),
            "scenarios": {"error": 0},
            "meta": {
                "step": 1,
                "description": "Σφάλμα παιδιών εκπαιδευτικών",
                "error": str(e)
            }
        }




# === HELPERS (STRICT placement of scenario column at a target Excel letter) ===
def _col_letter_to_index_strict(letter: str) -> int:
    """Μετατρέπει γράμμα στήλης Excel σε 0-based index (A->0, K->10)."""
    letter = str(letter).strip().upper()
    v = 0
    for ch in letter:
        if "A" <= ch <= "Z":
            v = v * 26 + (ord(ch) - ord("A") + 1)
    return max(0, v - 1)

def _reorder_with_padding(df: pd.DataFrame, scen_col: str, target_letter: str,
                          base_columns: list = None, pad_prefix: str = "__PAD_K") -> pd.DataFrame:
    """Επιστρέφει νέο DataFrame όπου η στήλη σεναρίου είναι ΑΥΣΤΗΡΑ στη target_letter.
Προσθέτει κενές στήλες padding αν χρειαστεί ώστε να υπάρχει πάντα η target_letter."""
    if base_columns is None:
        base_columns = ['ΟΝΟΜΑ','ΦΥΛΟ','ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ','ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ']

    df = df.copy()
    # Συγκεντρώνουμε όλες τις στήλες σεναρίων για να αφαιρέσουμε τις υπόλοιπες
    scenario_cols_all = [c for c in df.columns if str(c).startswith("ΒΗΜΑ1_ΣΕΝΑΡΙΟ_")]
    base_cols_present = [c for c in base_columns if c in df.columns]
    non_scen = [c for c in df.columns if c not in scenario_cols_all]
    remainder = [c for c in non_scen if c not in base_cols_present]

    ordered = base_cols_present + [c for c in remainder if c != scen_col]

    target_idx = _col_letter_to_index_strict(target_letter)
    # Padding μέχρι να υπάρχει τουλάχιστον target_idx θέσεις πριν την εισαγωγή
    pad_count = 0
    while len(ordered) < target_idx:
        pad_count += 1
        pad_name = f"{pad_prefix}{pad_count:02d}"
        while pad_name in df.columns or pad_name in ordered:
            pad_name += "_"
        df[pad_name] = ""
        ordered.append(pad_name)

    # Εισαγωγή στήλης σεναρίου ακριβώς στη θέση target_idx
    insert_idx = min(max(0, target_idx), len(ordered))
    ordered.insert(insert_idx, scen_col)

    return df[ordered]

def _hide_padding_columns(xlsx_path: str, pad_prefix: str = "__PAD_K"):
    """Κρύβει στο Excel όσες στήλες έχουν header που ξεκινά με pad_prefix."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    wb = load_workbook(xlsx_path)
    for ws in wb.worksheets:
        headers = [cell.value if cell.value is not None else "" for cell in ws[1]]
        for idx, name in enumerate(headers, start=1):
            if isinstance(name, str) and name.startswith(pad_prefix):
                ws.column_dimensions[get_column_letter(idx)].hidden = True
    wb.save(xlsx_path)
# === NEW (v2): Save Step-1 results with one sheet per scenario and place the scenario column at a target Excel letter ===
def save_immutable_step1_results_separate(df_with_step1: pd.DataFrame, results: Step1Results,
                                          output_file: str = "ΒΗΜΑ1_IMΜUTABLE_PER_ΣΕΝΑΡΙΟ.xlsx",
                                          include_input_sheet: bool = True,
                                          base_columns: list = None,
                                          scenario_col_letter: str = "K") -> None:
    """
    Αποθήκευση Βήματος 1 σε Excel με ΞΕΧΩΡΙΣΤΟ φύλλο για ΚΑΘΕ σενάριο και
    τοποθέτηση της στήλης του σεναρίου στη θέση της στήλης `scenario_col_letter` (π.χ. 'L').
    
    - Φύλλο 'ΣΕΝΑΡΙΑ_SUMMARY' με περιλήψεις.
    - Προαιρετικά φύλλο 'INPUT_DATA' με το πλήρες df.
    - Για κάθε σενάριο k: φύλλο 'ΣΕΝΑΡΙΟ_{k}' με reordered στήλες ώστε η στήλη του σεναρίου να είναι στη στήλη `scenario_col_letter`.
    
    Args:
        df_with_step1: DataFrame με τις στήλες ΒΗΜΑ1_ΣΕΝΑΡΙΟ_X
        results: Step1Results object
        output_file: Όνομα αρχείου εξόδου (.xlsx)
        include_input_sheet: Αν θα γραφτεί και φύλλο 'INPUT_DATA'
        base_columns: Ποιες βασικές στήλες να προηγούνται (default: ['ΟΝΟΜΑ','ΦΥΛΟ','ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ','ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ'])
        scenario_col_letter: Π.χ. 'L' => 12η στήλη (0-based index = 11)
    """
    if base_columns is None:
        base_columns = ['ΟΝΟΜΑ', 'ΦΥΛΟ', 'ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ', 'ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ']

    scenario_cols_all = [c for c in df_with_step1.columns if str(c).startswith("ΒΗΜΑ1_ΣΕΝΑΡΙΟ_")]

    def _col_letter_to_index(letter: str) -> int:
        letter = str(letter).strip().upper()
        val = 0
        for ch in letter:
            if not ("A" <= ch <= "Z"):
                continue
            val = val * 26 + (ord(ch) - ord("A") + 1)
        return max(0, val - 1)  # 0-based

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # Summary sheet
        summary_rows = []
        for s in results.scenarios:
            summary_rows.append({
                "ID": s.id,
                "Στήλη": s.column_name,
                "Περιγραφή": getattr(s, "description", ""),
                "Σπασμένες_Φιλίες": getattr(s, "broken_friendships", None),
                "Αναθέσεις": len(getattr(s, "assignments", [])),
                "Κλείδωμα": "IMMUTABLE"
            })
        pd.DataFrame(summary_rows).to_excel(writer, index=False, sheet_name="ΣΕΝΑΡΙΑ_SUMMARY")

        # Optional input data
        if include_input_sheet:
            df_with_step1.to_excel(writer, index=False, sheet_name="INPUT_DATA")

        # One sheet per scenario with target placement
        target_idx = _col_letter_to_index(scenario_col_letter)
        for s in results.scenarios:
            scen_col = s.column_name
            # Build column order: base columns first (in given order if they exist)
            base_cols_present = [c for c in base_columns if c in df_with_step1.columns]
            # All non-scenario columns
            non_scenario_cols = [c for c in df_with_step1.columns if c not in scenario_cols_all]
            # Move base cols to the very front (and keep their order)
            remainder = [c for c in non_scenario_cols if c not in base_cols_present]
            ordered = base_cols_present + [c for c in remainder if c != scen_col]
            # Insert the scenario column at the desired index (clamped)
            idx = min(max(0, target_idx), len(ordered))
            if scen_col in df_with_step1.columns:
                ordered.insert(idx, scen_col)
            # Write sheet
            df_s = df_with_step1[ordered].copy()
            df_s.to_excel(writer, index=False, sheet_name=f"ΣΕΝΑΡΙΟ_{s.id}")



# === NEW: Export ONLY one sheet per scenario (no summary, no input sheet) ===
def export_step1_per_scenario_only(df_with_step1: pd.DataFrame, results: Step1Results,
                                   output_file: str = "VIMA1_PER_SCENARIO_ONLY.xlsx",
                                   base_columns: list = None,
                                   scenario_col_letter: str = "K") -> None:
    """
    Γράφει ΜΟΝΟ ένα φύλλο ανά σενάριο (ΣΕΝΑΡΙΟ_{k}) και ΤΙΠΟΤΑ άλλο.
    Η στήλη του σεναρίου τοποθετείται στη στήλη `scenario_col_letter` (π.χ. 'L').
    """
    if base_columns is None:
        base_columns = ['ΟΝΟΜΑ','ΦΥΛΟ','ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ','ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ']

    # scenario columns
    scenario_cols_all = [c for c in df_with_step1.columns if str(c).startswith("ΒΗΜΑ1_ΣΕΝΑΡΙΟ_")]

    def _col_letter_to_index(letter: str) -> int:
        letter = str(letter).strip().upper()
        v=0
        for ch in letter:
            if "A"<=ch<="Z":
                v = v*26 + (ord(ch)-ord("A")+1)
        return max(0, v-1)

    target_idx = _col_letter_to_index(scenario_col_letter)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for s in results.scenarios:
            scen_col = s.column_name
            base_cols_present = [c for c in base_columns if c in df_with_step1.columns]
            non_scen = [c for c in df_with_step1.columns if c not in scenario_cols_all]
            remainder = [c for c in non_scen if c not in base_cols_present]
            ordered = base_cols_present + [c for c in remainder if c != scen_col]
            idx = min(max(0, target_idx), len(ordered))
            if scen_col in df_with_step1.columns:
                ordered.insert(idx, scen_col)
            df_with_step1[ordered].to_excel(writer, index=False, sheet_name=f"ΣΕΝΑΡΙΟ_{s.id}")



# === STRICT VARIANTS (appended) ===
def save_immutable_step1_results_separate_STRICT(df_with_step1: pd.DataFrame, results: Step1Results,
                                                 output_file: str = "ΒΗΜΑ1_IMΜUTABLE_PER_ΣΕΝΑΡΙΟ.xlsx",
                                                 include_input_sheet: bool = True,
                                                 base_columns: list = None,
                                                 scenario_col_letter: str = "K",
                                                 hide_padding: bool = True) -> None:
    if base_columns is None:
        base_columns = ['ΟΝΟΜΑ','ΦΥΛΟ','ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ','ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ']

    summary_rows = []
    for s in results.scenarios:
        summary_rows.append({
            "ID": s.id,
            "Στήλη": s.column_name,
            "Περιγραφή": getattr(s, "description", ""),
            "Σπασμένες_Φιλίες": getattr(s, "broken_friendships", None),
            "Αναθέσεις": len(getattr(s, "assignments", [])),
            "Κλείδωμα": "IMMUTABLE"
        })
    summary_df = pd.DataFrame(summary_rows)

    tmp_out = output_file + (".tmp.xlsx" if hide_padding else "")
    with pd.ExcelWriter(tmp_out if hide_padding else output_file, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="ΣΕΝΑΡΙΑ_SUMMARY")
        if include_input_sheet:
            df_with_step1.to_excel(writer, index=False, sheet_name="INPUT_DATA")
        for s in results.scenarios:
            scen_col = s.column_name
            df_s = _reorder_with_padding(df_with_step1, scen_col, scenario_col_letter, base_columns)
            df_s.to_excel(writer, index=False, sheet_name=f"ΣΕΝΑΡΙΟ_{s.id}")

    if hide_padding:
        _hide_padding_columns(tmp_out)
        Path(tmp_out).replace(output_file)


def export_step1_per_scenario_only_STRICT(df_with_step1: pd.DataFrame, results: Step1Results,
                                          output_file: str = "VIMA1_PER_SCENARIO_ONLY.xlsx",
                                          base_columns: list = None,
                                          scenario_col_letter: str = "K",
                                          hide_padding: bool = True) -> None:
    if base_columns is None:
        base_columns = ['ΟΝΟΜΑ','ΦΥΛΟ','ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ','ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ']

    tmp_out = output_file + (".tmp.xlsx" if hide_padding else "")
    with pd.ExcelWriter(tmp_out if hide_padding else output_file, engine="openpyxl") as writer:
        for s in results.scenarios:
            scen_col = s.column_name
            df_s = _reorder_with_padding(df_with_step1, scen_col, scenario_col_letter, base_columns)
            df_s.to_excel(writer, index=False, sheet_name=f"ΣΕΝΑΡΙΟ_{s.id}")

    if hide_padding:
        _hide_padding_columns(tmp_out)
        Path(tmp_out).replace(output_file)

# Override old function names to point to STRICT implementations
save_immutable_step1_results_separate = save_immutable_step1_results_separate_STRICT
export_step1_per_scenario_only = export_step1_per_scenario_only_STRICT
