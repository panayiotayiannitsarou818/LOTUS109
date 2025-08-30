# -*- coding: utf-8 -*-
"""
Enhanced VIMA6 Export Module
Εξαγωγή Excel με όλα τα βήματα και σενάρια σε μορφή ΑΝΑΛΥΤΙΚΑ_ΒΗΜΑΤΑ
"""

import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re
from typing import Dict, List, Any, Optional


def extract_all_scenario_data(final_results: Dict[str, Dict]) -> pd.DataFrame:
    """
    Εξαγωγή δεδομένων από όλα τα σενάρια για VIMA6 format
    
    Args:
        final_results: Dict με αποτελέσματα από όλα τα βήματα
        
    Returns:
        DataFrame με στήλες: ΟΝΟΜΑ, ΒΗΜΑ1_ΣΕΝΑΡΙΟ_1, ΒΗΜΑ2_ΣΕΝΑΡΙΟ_1, ..., ΒΗΜΑ6_ΣΕΝΑΡΙΟ_X
    """
    if not final_results:
        return pd.DataFrame()
    
    # Βρες το πρώτο DataFrame για να πάρεις τα ονόματα
    first_scenario = list(final_results.values())[0]
    base_df = first_scenario.get('df')
    if base_df is None or 'ΟΝΟΜΑ' not in base_df.columns:
        return pd.DataFrame()
    
    # Αρχικοποίηση με τα ονόματα
    result_df = pd.DataFrame()
    result_df['ΟΝΟΜΑ'] = base_df['ΟΝΟΜΑ'].copy()
    
    # Για κάθε σενάριο, βρες όλες τις ΒΗΜΑ στήλες
    scenario_count = 0
    for scenario_name, scenario_data in final_results.items():
        scenario_count += 1
        if scenario_count > 5:  # Μέχρι 5 σενάρια
            break
            
        df = scenario_data.get('df')
        if df is None:
            continue
            
        # Εξαγωγή αριθμού σεναρίου από το όνομα
        scenario_match = re.search(r'(\d+)$', scenario_name)
        scenario_num = scenario_match.group(1) if scenario_match else str(scenario_count)
        
        # Βρες όλες τις ΒΗΜΑ στήλες για αυτό το σενάριο
        step_columns = []
        for col in df.columns:
            if re.match(r'^ΒΗΜΑ\d+_', str(col)):
                step_columns.append(col)
        
        # Ταξινόμηση των στηλών κατά αριθμό βήματος
        step_columns.sort(key=lambda x: int(re.search(r'ΒΗΜΑ(\d+)', x).group(1)) if re.search(r'ΒΗΜΑ(\d+)', x) else 0)
        
        # Προσθήκη στηλών στο αποτέλεσμα
        for step_col in step_columns:
            step_match = re.search(r'ΒΗΜΑ(\d+)', step_col)
            if step_match:
                step_num = step_match.group(1)
                new_col_name = f"ΒΗΜΑ{step_num}_ΣΕΝΑΡΙΟ_{scenario_num}"
                
                if new_col_name not in result_df.columns:
                    result_df[new_col_name] = df[step_col] if step_col in df.columns else None
        
        # Αν υπάρχει τελική στήλη ανάθεσης, πρόσθεσέ την ως ΒΗΜΑ6
        final_column = scenario_data.get('final_column')
        if final_column and final_column in df.columns:
            final_col_name = f"ΒΗΜΑ6_ΣΕΝΑΡΙΟ_{scenario_num}"
            if final_col_name not in result_df.columns:
                result_df[final_col_name] = df[final_column]
    
    return result_df


def calculate_scores_summary(final_results: Dict[str, Dict]) -> pd.DataFrame:
    """
    Υπολογισμός βαθμολογιών από βήμα 7 για όλα τα σενάρια
    """
    scores_data = []
    
    for scenario_name, scenario_data in final_results.items():
        score_info = scenario_data.get('final_score', {})
        
        if score_info:
            scores_data.append({
                'ΣΕΝΑΡΙΟ': scenario_name,
                'ΣΥΝΟΛΙΚΟ_SCORE': score_info.get('total_score', 0),
                'ΔΙΑΦΟΡΑ_ΠΛΗΘΥΣΜΟΥ': score_info.get('diff_population', 0),
                'ΔΙΑΦΟΡΑ_ΦΥΛΟΥ': score_info.get('diff_gender', 0),
                'ΔΙΑΦΟΡΑ_ΓΝΩΣΗΣ': score_info.get('diff_greek', 0),
                'ΣΠΑΣΜΕΝΕΣ_ΦΙΛΙΕΣ': score_info.get('broken_friendships', 0),
                'RANK': 0  # Θα υπολογιστεί παρακάτω
            })
    
    # Δημιουργία DataFrame και ταξινόμηση
    scores_df = pd.DataFrame(scores_data)
    if not scores_df.empty:
        # Ταξινόμηση κατά συνολικό score (χαμηλότερο = καλύτερο)
        scores_df = scores_df.sort_values('ΣΥΝΟΛΙΚΟ_SCORE')
        scores_df['RANK'] = range(1, len(scores_df) + 1)
        scores_df = scores_df.reset_index(drop=True)
    
    return scores_df


def export_vima6_complete(final_results: Dict[str, Dict], filename: str = "ΑΝΑΛΥΤΙΚΑ_ΒΗΜΑΤΑ_VIMA6.xlsx") -> BytesIO:
    """
    Εξαγωγή πλήρους VIMA6 Excel με όλα τα σενάρια και βαθμολογίες
    
    Args:
        final_results: Αποτελέσματα από όλα τα βήματα
        filename: Όνομα αρχείου εξαγωγής
        
    Returns:
        BytesIO buffer με Excel αρχείο
    """
    # Εξαγωγή δεδομένων βημάτων
    vima_df = extract_all_scenario_data(final_results)
    scores_df = calculate_scores_summary(final_results)
    
    # Δημιουργία Excel
    buffer = BytesIO()
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # Sheet 1: Αναλυτικά Βήματα
        if not vima_df.empty:
            vima_df.to_excel(writer, sheet_name='ΑΝΑΛΥΤΙΚΑ_ΒΗΜΑΤΑ', index=False)
        
        # Sheet 2: Βαθμολογίες
        if not scores_df.empty:
            scores_df.to_excel(writer, sheet_name='ΒΑΘΜΟΛΟΓΙΕΣ', index=False)
    
    buffer.seek(0)
    
    # Φορμάτισμα με openpyxl
    wb = load_workbook(buffer)
    
    # Φορμάτισμα sheet ΑΝΑΛΥΤΙΚΑ_ΒΗΜΑΤΑ
    if 'ΑΝΑΛΥΤΙΚΑ_ΒΗΜΑΤΑ' in wb.sheetnames and not vima_df.empty:
        ws = wb['ΑΝΑΛΥΤΙΚΑ_ΒΗΜΑΤΑ']
        format_vima_sheet(ws)
    
    # Φορμάτισμα sheet ΒΑΘΜΟΛΟΓΙΕΣ
    if 'ΒΑΘΜΟΛΟΓΙΕΣ' in wb.sheetnames and not scores_df.empty:
        ws_scores = wb['ΒΑΘΜΟΛΟΓΙΕΣ']
        format_scores_sheet(ws_scores)
    
    # Αποθήκευση
    formatted_buffer = BytesIO()
    wb.save(formatted_buffer)
    formatted_buffer.seek(0)
    formatted_buffer.name = filename
    
    return formatted_buffer


def format_vima_sheet(ws) -> None:
    """Φορμάτισμα του sheet ΑΝΑΛΥΤΙΚΑ_ΒΗΜΑΤΑ"""
    # Παγώματα στήλης Α
    ws.freeze_panes = 'B2'
    
    # Χρώματα και fonts για headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Φορμάτισμα πρώτης γραμμής (headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
    
    # Κεντράρισμα δεδομένων (εκτός από στήλη ονομάτων)
    for col in range(2, ws.max_column + 1):
        for row in range(2, ws.max_row + 1):
            ws.cell(row, col).alignment = center_alignment
    
    # Auto-adjust column widths
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_length = 0
        for row in range(1, min(ws.max_row + 1, 100)):  # Check first 100 rows for performance
            cell_value = str(ws.cell(row, col).value or "")
            max_length = max(max_length, len(cell_value))
        
        # Set width (με όρια)
        width = min(max(max_length + 2, 10), 25)
        ws.column_dimensions[letter].width = width


def format_scores_sheet(ws) -> None:
    """Φορμάτισμα του sheet ΒΑΘΜΟΛΟΓΙΕΣ"""
    # Χρώματα για βαθμολογίες
    header_fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
    
    # Κεντράρισμα όλων των δεδομένων
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).alignment = center_alignment
    
    # Χρωματισμός του καλύτερου σεναρίου (RANK = 1)
    best_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    for row in range(2, ws.max_row + 1):
        rank_cell = ws.cell(row, ws.max_column)  # Τελευταία στήλη = RANK
        if rank_cell.value == 1:
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).fill = best_fill
    
    # Auto-adjust column widths
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_length = max(len(str(ws.cell(1, col).value or "")), 15)
        ws.column_dimensions[letter].width = max_length + 2


# Helper function για integration με υπάρχον κώδικα
def create_vima6_download_button(final_results: Dict[str, Dict], streamlit_instance) -> None:
    """
    Δημιουργία download button για VIMA6 στο Streamlit
    
    Args:
        final_results: Αποτελέσματα από final steps
        streamlit_instance: Streamlit instance (st)
    """
    if not final_results:
        streamlit_instance.warning("Δεν υπάρχουν αποτελέσματα για εξαγωγή")
        return
    
    try:
        vima6_buffer = export_vima6_complete(final_results)
        
        streamlit_instance.download_button(
            label="📋 Αναλυτικά Βήματα (VIMA6) + Βαθμολογίες",
            data=vima6_buffer.getvalue(),
            file_name=vima6_buffer.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Εξαγωγή πλήρων αναλυτικών βημάτων με εξέλιξη ανά σενάριο και βαθμολογίες"
        )
        
    except Exception as e:
        streamlit_instance.error(f"Σφάλμα στη δημιουργία VIMA6: {e}")
