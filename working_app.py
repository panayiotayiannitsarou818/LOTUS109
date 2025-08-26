# === VIMA6 Excel builder (added 2025-08-25) ===
import re
import traceback
from io import BytesIO
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

def _num(s, pat):
    m = re.search(pat, s)
    return int(m.group(1)) if m else 999

def _ordered_steps(keys):
    # Sort by step number in "ΒΗΜΑx"
    return sorted(keys, key=lambda k: _num(k, r"ΒΗΜΑ(\d+)"))

def _compose_new_plus_prev(step_series_list):
    """
    Build **cumulative** display columns:
    - col[0] = step0 (non-empty values only)
    - For i>0: col[i] starts as col[i-1] (carry-forward), then overlay non-empty values from step i.
    This matches the requirement:
      Βήμα 2 = (ό,τι υπάρχει στο Βήμα 1) + (οι νέες τοποθετήσεις του Βήματος 2)
      Βήμα 3 = (ό,τι υπάρχει στο Βήμα 2) + (οι νέες του Βήματος 3), κ.ο.κ.
    """
    import pandas as pd
    out_cols = []
    prev_col = None
    for i, curr in enumerate(step_series_list):
        curr = curr.astype(str).replace({None: ""}).fillna("")
        if i == 0:
            col = pd.Series([""] * len(curr), index=curr.index, dtype=object)
            non_empty = curr != ""
            col[non_empty] = curr[non_empty]
        else:
            # start from previous display column (carry-forward cumulative state)
            col = prev_col.copy()
            non_empty = curr != ""
            # overlay only non-empty values of current step
            col[non_empty] = curr[non_empty]
        out_cols.append(col)
        prev_col = col
    return out_cols

def build_vima6_excel_bytes(base_df, detailed_steps, step7_scores=None):
    """
    Creates Excel bytes with:
      - Sheet 'ΑΝΑΛΥΤΙΚΑ_ΒΗΜΑΤΑ' (cumulative per step): 
        Βήμα N = (ό,τι υπάρχει στο Βήμα N-1) + (οι μη-κενές τιμές του Βήματος N)
      - Optional sheet 'VIMA7_SCORE_ΣΥΝΟΨΗ'
    Supports input shapes:
      (A) working_app style: {"ΒΗΜΑ1_ΣΕΝΑΡΙΟ_1": df, ...}
      (B) streamlit_app style: { scen: {"data": {...}, "step7_score": int, ...}, ...}
    """
    import pandas as pd
    from io import BytesIO
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    out = base_df[['ΟΝΟΜΑ']].copy()

    if detailed_steps and all(isinstance(k, str) for k in detailed_steps.keys()):
        # --- working_app style ---
        scen_to_steps = {}
        for key, df_k in detailed_steps.items():
            m = re.search(r"_ΣΕΝΑΡΙΟ_(\d+)$", key)
            scen = int(m.group(1)) if m else 1
            scen_to_steps.setdefault(scen, {})[key] = df_k

        for scen, step_map in sorted(scen_to_steps.items()):
            ordered = _ordered_steps(step_map.keys())
            series_list = []
            for step_key in ordered:
                df_k = step_map[step_key]
                if step_key in df_k.columns:
                    s = df_k[step_key]
                else:
                    s = pd.Series([""] * len(base_df), index=base_df.index, dtype=object)
                series_list.append(s.astype(str).replace({None: ""}).fillna(""))
            cols = _compose_new_plus_prev(series_list)  # cumulative
            for step_key, col in zip(ordered, cols):
                out[step_key] = col
    else:
        # --- streamlit_app style ---
        for scen_num, scen_data in sorted(detailed_steps.items()):
            view = None
            try:
                view = build_analytics_view_upto6_with_score(base_df, scen_data, scen_num)
            except Exception:
                data = scen_data.get('data', {})
                step_keys = sorted(
                    [k for k in data.keys() if isinstance(k, str) and k.startswith("ΒΗΜΑ")],
                    key=lambda k: _num(k, r"ΒΗΜΑ(\d+)")
                )
                if step_keys:
                    series_list = []
                    for k in step_keys:
                        s = pd.Series(data.get(k, [""] * len(base_df)))
                        series_list.append(s.astype(str).replace({None: ""}).fillna(""))
                    cols = _compose_new_plus_prev(series_list)
                    view = pd.DataFrame({"ΟΝΟΜΑ": base_df["ΟΝΟΜΑ"]})
                    for k, c in zip(step_keys, cols):
                        view[k] = c
                else:
                    view = pd.DataFrame({"ΟΝΟΜΑ": base_df["ΟΝΟΜΑ"]})
            for c in [c for c in view.columns if c != 'ΟΝΟΜΑ']:
                out[c] = view[c]

    bio = BytesIO()
    with pd.ExcelWriter(bio, engine='openpyxl') as writer:
        out.to_excel(writer, sheet_name="ΑΝΑΛΥΤΙΚΑ_ΒΗΜΑΤΑ", index=False)
        ws = writer.sheets["ΑΝΑΛΥΤΙΚΑ_ΒΗΜΑΤΑ"]
        ws.freeze_panes = "B2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
        max_col = ws.max_column
        max_row = ws.max_row
        for col_idx in range(2, max_col + 1):
            for row_idx in range(2, max_row + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=col_idx, max_col=col_idx):
                val = row[0].value
                max_len = max(max_len, len(str(val)) if val is not None else 0)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        # Optional VIMA7 scores
        scores_rows = []
        if detailed_steps and not all(isinstance(k, str) for k in detailed_steps.keys()):
            for scen_num, scen_data in sorted(detailed_steps.items()):
                sc = scen_data.get("step7_score")
                if sc is not None:
                    scores_rows.append({"ΣΕΝΑΡΙΟ": f"ΣΕΝΑΡΙΟ_{scen_num}", "ΤΕΛΙΚΗ_ΒΑΘΜΟΛΟΓΙΑ_ΣΕΝΑΡΙΟΥ": sc})
        if step7_scores:
            for scen_num, sc in sorted(step7_scores.items()):
                scores_rows.append({"ΣΕΝΑΡΙΟ": f"ΣΕΝΑΡΙΟ_{scen_num}", "ΤΕΛΙΚΗ_ΒΑΘΜΟΛΟΓΙΑ_ΣΕΝΑΡΙΟΥ": sc})
        if scores_rows:
            pd.DataFrame(scores_rows).to_excel(writer, sheet_name="VIMA7_SCORE_ΣΥΝΟΨΗ", index=False)
            ws2 = writer.sheets["VIMA7_SCORE_ΣΥΝΟΨΗ"]
            ws2.freeze_panes = "A2"
            for cell in ws2[1]:
                cell.font = Font(bold=True)
            for col_idx in range(1, ws2.max_column + 1):
                from openpyxl.utils import get_column_letter as _gcl2
                col_letter = _gcl2(col_idx)
                max_len = 0
                for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=col_idx, max_col=col_idx):
                    val = row[0].value
                    max_len = max(max_len, len(str(val)) if val is not None else 0)
                ws2.column_dimensions[col_letter].width = min(max_len + 2, 60)

    bio.seek(0)
    return bio.getvalue()

# -*- coding: utf-8 -*-
"""
ΛΕΙΤΟΥΡΓΙΚΗ ΕΚΔΟΣΗ - Με βήματα ανάθεσης αλλά χωρίς περίπλοκα γραφήματα
ΠΡΟΣΤΕΘΗΚΕ: Κουμπί για αναλυτικά βήματα (VIMA6 format)
"""

import streamlit as st
import pandas as pd
import numpy as np
import zipfile
import io
from typing import Dict, List, Tuple, Any
import traceback

# Import των modules που χρειάζονται
try:
    from statistics_generator import generate_statistics_table, export_statistics_to_excel
    STATS_AVAILABLE = True
except ImportError:
    STATS_AVAILABLE = False
    st.warning("⚠️ Module statistics_generator δεν βρέθηκε. Τα στατιστικά θα είναι περιορισμένα.")

# Streamlit configuration
st.set_page_config(
    page_title="Σύστημα Ανάθεσης Μαθητών",
    page_icon="🎓",
    layout="wide"
)

def init_session_state():
    """Αρχικοποίηση session state"""
    if 'data' not in st.session_state:
        st.session_state.data = None
    if 'current_step' not in st.session_state:
        st.session_state.current_step = 0
    if 'results' not in st.session_state:
        st.session_state.results = {}
    # ΠΡΟΣΤΕΘΗΚΕ: Για αποθήκευση ενδιάμεσων βημάτων
    if 'detailed_steps' not in st.session_state:
        st.session_state.detailed_steps = {}

def safe_load_data(uploaded_file):
    """Ασφαλής φόρτωση και κανονικοποίηση δεδομένων"""
    try:
        if uploaded_file.name.endswith('.xlsx'):
            df = pd.read_excel(uploaded_file)
        elif uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file)
        else:
            return None, "Μη υποστηριζόμενο format αρχείου"
        
        # Debug: Εμφάνιση αρχικών στηλών
        st.write("**DEBUG - Αρχικές στήλες:**", list(df.columns))
        st.write("**DEBUG - Πρώτες 3 γραμμές:**")
        st.dataframe(df.head(3))
        
        # Κανονικοποίηση στηλών με πιο επιθετική αναζήτηση
        rename_map = {}
        for col in df.columns:
            col_str = str(col).strip().upper()
            col_clean = col_str.replace(' ', '_').replace('-', '_')
            
            # Ονόματα
            if any(x in col_clean for x in ['ΟΝΟΜΑ', 'ONOMA', 'NAME', 'ΜΑΘΗΤΗΣ', 'ΜΑΘΗΤΡΙΑ', 'STUDENT']):
                rename_map[col] = 'ΟΝΟΜΑ'
            # Φύλο    
            elif any(x in col_clean for x in ['ΦΥΛΟ', 'FYLO', 'GENDER', 'SEX']):
                rename_map[col] = 'ΦΥΛΟ'
            # Γνώση Ελληνικών - περισσότερες παραλλαγές
            elif any(pattern in col_clean for pattern in ['ΓΝΩΣΗ', 'ΓΝΩΣΕΙΣ', 'ΕΛΛΗΝΙΚ', 'ELLINIK', 'GREEK']):
                rename_map[col] = 'ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ'
            # Παιδιά εκπαιδευτικών - περισσότερες παραλλαγές  
            elif any(pattern in col_clean for pattern in ['ΠΑΙΔΙ', 'PAIDI', 'ΕΚΠΑΙΔΕΥΤΙΚ', 'EKPEDEFTIK', 'TEACHER', 'ΔΑΣΚΑΛ']):
                rename_map[col] = 'ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ'
            # Φίλοι
            elif any(x in col_clean for x in ['ΦΙΛΟΙ', 'FILOI', 'FRIEND']):
                rename_map[col] = 'ΦΙΛΟΙ'
            # Ζωηρότητα
            elif any(x in col_clean for x in ['ΖΩΗΡ', 'ZOIR', 'ACTIVE', 'ENERGY']):
                rename_map[col] = 'ΖΩΗΡΟΣ'
            # Ιδιαιτερότητες
            elif any(x in col_clean for x in ['ΙΔΙΑΙΤΕΡΟΤΗΤ', 'IDIETEROTIT', 'SPECIAL']):
                rename_map[col] = 'ΙΔΙΑΙΤΕΡΟΤΗΤΑ'
            # Συγκρούσεις
            elif any(x in col_clean for x in ['ΣΥΓΚΡΟΥΣ', 'SYGKROUS', 'CONFLICT']):
                rename_map[col] = 'ΣΥΓΚΡΟΥΣΗ'
        
        st.write("**DEBUG - Rename map:**", rename_map)
        
        if rename_map:
            df = df.rename(columns=rename_map)
            
        st.write("**DEBUG - Μετά rename:**", list(df.columns))
        
        # Κανονικοποίηση τιμών με περισσότερες παραλλαγές
        if 'ΦΥΛΟ' in df.columns:
            # Κανονικοποίηση φύλου
            df['ΦΥΛΟ'] = df['ΦΥΛΟ'].astype(str).str.strip().str.upper()
            gender_map = {
                'Α': 'Α', 'ΑΓΟΡΙ': 'Α', 'ΑΓΟΡΙΟΥ': 'Α', 'BOY': 'Α', 'MALE': 'Α', 'M': 'Α',
                'Κ': 'Κ', 'ΚΟΡΙΤΣΙ': 'Κ', 'ΚΟΡΙΤΣΙΟΥ': 'Κ', 'GIRL': 'Κ', 'FEMALE': 'Κ', 'F': 'Κ'
            }
            df['ΦΥΛΟ'] = df['ΦΥΛΟ'].map(gender_map).fillna('Α')
            st.write("**DEBUG - Φύλο unique values:**", df['ΦΥΛΟ'].unique())
        
        # Κανονικοποίηση boolean στηλών
        bool_columns = ['ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ', 'ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ', 'ΖΩΗΡΟΣ', 'ΙΔΙΑΙΤΕΡΟΤΗΤΑ']
        for col in bool_columns:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip().str.upper()
                bool_map = {
                    'Ν': 'Ν', 'ΝΑΙ': 'Ν', 'YES': 'Ν', 'Y': 'Ν', '1': 'Ν', 'TRUE': 'Ν', 'T': 'Ν',
                    'Ο': 'Ο', 'ΟΧΙ': 'Ο', 'NO': 'Ο', 'N': 'Ο', '0': 'Ο', 'FALSE': 'Ο', 'F': 'Ο'
                }
                df[col] = df[col].map(bool_map).fillna('Ο')
                st.write(f"**DEBUG - {col} unique values:**", df[col].unique())
        
        return df, None
    except Exception as e:
        return None, f"Σφάλμα φόρτωσης: {str(e)}"

def display_basic_info(df):
    """Βασικές πληροφορίες με debug"""
    st.subheader("📊 Βασικές Πληροφορίες")
    
    total_students = len(df)
    
    # Υπολογισμός με debug
    boys_count = 0
    girls_count = 0
    teachers_count = 0
    greek_count = 0
    
    if 'ΦΥΛΟ' in df.columns:
        boys_count = len(df[df['ΦΥΛΟ'] == 'Α'])
        girls_count = len(df[df['ΦΥΛΟ'] == 'Κ'])
        st.write(f"**DEBUG - ΦΥΛΟ:** Α={boys_count}, Κ={girls_count}, Unique: {df['ΦΥΛΟ'].unique()}")
    
    if 'ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ' in df.columns:
        teachers_count = len(df[df['ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ'] == 'Ν'])
        teachers_list = df[df['ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ'] == 'Ν']['ΟΝΟΜΑ'].tolist() if 'ΟΝΟΜΑ' in df.columns else []
        st.write(f"**DEBUG - ΠΑΙΔΙΈ ΕΚΠΑΙΔΕΥΤΙΚΏΝ:** {teachers_count} άτομα")
        st.write(f"**DEBUG - Unique values:** {df['ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ'].unique()}")
        if teachers_list:
            st.write(f"**DEBUG - Παιδιά εκπαιδευτικών:** {teachers_list}")
    
    if 'ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ' in df.columns:
        greek_count = len(df[df['ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ'] == 'Ν'])
        st.write(f"**DEBUG - ΓΝΩΣΗ ΕΛΛΗΝΙΚΩΝ:** Ν={greek_count}, Unique: {df['ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ'].unique()}")
    
    # Εμφάνιση μετρικών
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Συνολικοί Μαθητές", total_students)
    with col2:
        st.metric("Αγόρια", boys_count)
    with col3:
        st.metric("Κορίτσια", girls_count)
    with col4:
        st.metric("Παιδιά Εκπαιδευτικών", teachers_count)

def display_scenario_stats(df, scenario_col, scenario_name):
    """Εμφάνιση στατιστικών σεναρίου με βελτιωμένη λογική"""
    try:
        if scenario_col not in df.columns:
            st.warning(f"Η στήλη {scenario_col} δεν βρέθηκε")
            return
            
        df_assigned = df[df[scenario_col].notna()].copy()
        if len(df_assigned) == 0:
            st.warning("Δεν βρέθηκαν τοποθετημένοι μαθητές")
            return
            
        st.subheader(f"📊 Στατιστικά {scenario_name}")
        
        # Χειροκίνητη δημιουργία στατιστικών - πιο αξιόπιστη
        stats_data = []
        for tmima in sorted(df_assigned[scenario_col].unique()):
            subset = df_assigned[df_assigned[scenario_col] == tmima]
            
            # Μετρήσεις με debug
            boys = len(subset[subset['ΦΥΛΟ'] == 'Α']) if 'ΦΥΛΟ' in subset.columns else 0
            girls = len(subset[subset['ΦΥΛΟ'] == 'Κ']) if 'ΦΥΛΟ' in subset.columns else 0
            
            # Υπολογισμός εκπαιδευτικών
            educators = 0
            if 'ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ' in subset.columns:
                educators = len(subset[subset['ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ'] == 'Ν'])
            
            # Υπολογισμός ζωηρών
            energetic = 0
            if 'ΖΩΗΡΟΣ' in subset.columns:
                energetic = len(subset[subset['ΖΩΗΡΟΣ'] == 'Ν'])
            
            # Υπολογισμός ιδιαιτεροτήτων
            special = 0
            if 'ΙΔΙΑΙΤΕΡΟΤΗΤΑ' in subset.columns:
                special = len(subset[subset['ΙΔΙΑΙΤΕΡΟΤΗΤΑ'] == 'Ν'])
            
            # Υπολογισμός γνώσης ελληνικών
            greek = 0
            if 'ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ' in subset.columns:
                greek = len(subset[subset['ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ'] == 'Ν'])
            
            total = len(subset)
            
            stats_data.append({
                'ΤΜΗΜΑ': tmima,
                'ΑΓΟΡΙΑ': boys,
                'ΚΟΡΙΤΣΙΑ': girls,
                'ΕΚΠΑΙΔΕΥΤΙΚΟΙ': educators,
                'ΖΩΗΡΟΙ': energetic,
                'ΙΔΙΑΙΤΕΡΟΤΗΤΑ': special,
                'ΓΝΩΣΗ ΕΛΛ.': greek,
                'ΣΥΝΟΛΟ': total
            })
        
        stats_df = pd.DataFrame(stats_data)
        st.dataframe(stats_df, use_container_width=True)
        
        # Debug info
        st.write("**DEBUG - Αναλυτικά ανά τμήμα:**")
        for tmima in sorted(df_assigned[scenario_col].unique()):
            subset = df_assigned[df_assigned[scenario_col] == tmima]
            st.write(f"**{tmima}:** {len(subset)} μαθητές")
            if 'ΟΝΟΜΑ' in subset.columns:
                names = subset['ΟΝΟΜΑ'].tolist()
                st.write(f"Ονόματα: {', '.join(names[:5])}{'...' if len(names) > 5 else ''}")
            
    except Exception as e:
        st.error(f"Σφάλμα στα στατιστικά: {e}")
        st.code(traceback.format_exc())

def run_simple_assignment(df):
    """Απλή ανάθεση με debug και βελτιωμένη λογική"""
    try:
        st.subheader("🚀 Εκτέλεση Απλής Ανάθεσης")
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        # Βήμα 1: Παιδιά εκπαιδευτικών
        status_text.text("Βήμα 1: Ανάθεση παιδιών εκπαιδευτικών...")
        progress_bar.progress(20)
        
        df_result = df.copy()
        df_result['ΤΜΗΜΑ'] = None
        
        # ΑΠΟΘΗΚΕΥΣΗ ΒΗΜΑ 1 (για αναλυτικά βήματα)
        df_step1 = df_result.copy()
        df_step1['ΒΗΜΑ1_ΣΕΝΑΡΙΟ_1'] = None
        
        # Debug: Έλεγχος παιδιών εκπαιδευτικών
        if 'ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ' in df_result.columns:
            teacher_kids = df_result[df_result['ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ'] == 'Ν'].index.tolist()
            st.write(f"**DEBUG - Παιδιά εκπαιδευτικών βρέθηκαν:** {len(teacher_kids)}")
            
            if teacher_kids and 'ΟΝΟΜΑ' in df_result.columns:
                teacher_names = df_result.loc[teacher_kids, 'ΟΝΟΜΑ'].tolist()
                st.write(f"**DEBUG - Ονόματα:** {teacher_names}")
            
            # Απλή κατανομή παιδιών εκπαιδευτικών
            for i, idx in enumerate(teacher_kids):
                tmima = 'Α1' if i % 2 == 0 else 'Α2'
                df_result.loc[idx, 'ΤΜΗΜΑ'] = tmima
                df_step1.loc[idx, 'ΒΗΜΑ1_ΣΕΝΑΡΙΟ_1'] = tmima  # Για αναλυτικά
                st.write(f"**DEBUG - Ανάθεση:** {df_result.loc[idx, 'ΟΝΟΜΑ'] if 'ΟΝΟΜΑ' in df_result.columns else idx} → {tmima}")
        else:
            st.warning("⚠️ Δεν βρέθηκε στήλη ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ")
        
        # ΑΠΟΘΗΚΕΥΣΗ ΓΙΑ ΑΝΑΛΥΤΙΚΑ ΒΗΜΑΤΑ
        st.session_state.detailed_steps['ΒΗΜΑ1_ΣΕΝΑΡΙΟ_1'] = df_step1.copy()
        
        progress_bar.progress(40)
        
        # Βήμα 2: Υπόλοιποι μαθητές
        status_text.text("Βήμα 2: Ανάθεση υπόλοιπων μαθητών...")
        
        remaining = df_result[df_result['ΤΜΗΜΑ'].isna()].index.tolist()
        st.write(f"**DEBUG - Υπόλοιποι μαθητές:** {len(remaining)}")
        
        # ΑΠΟΘΗΚΕΥΣΗ ΒΗΜΑ 2
        df_step2 = df_result.copy()
        df_step2['ΒΗΜΑ2_ΣΕΝΑΡΙΟ_1'] = df_step2['ΤΜΗΜΑ']
        
        if 'ΦΥΛΟ' in df_result.columns:
            # Διαχωρισμός ανά φύλο
            boys = [idx for idx in remaining if df_result.loc[idx, 'ΦΥΛΟ'] == 'Α']
            girls = [idx for idx in remaining if df_result.loc[idx, 'ΦΥΛΟ'] == 'Κ']
            
            st.write(f"**DEBUG - Αγόρια για ανάθεση:** {len(boys)}")
            st.write(f"**DEBUG - Κορίτσια για ανάθεση:** {len(girls)}")
            
            # Κατανομή αγοριών
            for i, idx in enumerate(boys):
                current_a1_boys = len(df_result[(df_result['ΤΜΗΜΑ'] == 'Α1') & (df_result['ΦΥΛΟ'] == 'Α')])
                current_a2_boys = len(df_result[(df_result['ΤΜΗΜΑ'] == 'Α2') & (df_result['ΦΥΛΟ'] == 'Α')])
                
                if current_a1_boys <= current_a2_boys:
                    df_result.loc[idx, 'ΤΜΗΜΑ'] = 'Α1'
                    df_step2.loc[idx, 'ΒΗΜΑ2_ΣΕΝΑΡΙΟ_1'] = 'Α1'
                else:
                    df_result.loc[idx, 'ΤΜΗΜΑ'] = 'Α2'
                    df_step2.loc[idx, 'ΒΗΜΑ2_ΣΕΝΑΡΙΟ_1'] = 'Α2'
            
            progress_bar.progress(60)
            
            # Κατανομή κοριτσιών
            for i, idx in enumerate(girls):
                current_a1_girls = len(df_result[(df_result['ΤΜΗΜΑ'] == 'Α1') & (df_result['ΦΥΛΟ'] == 'Κ')])
                current_a2_girls = len(df_result[(df_result['ΤΜΗΜΑ'] == 'Α2') & (df_result['ΦΥΛΟ'] == 'Κ')])
                
                if current_a1_girls <= current_a2_girls:
                    df_result.loc[idx, 'ΤΜΗΜΑ'] = 'Α1'
                    df_step2.loc[idx, 'ΒΗΜΑ2_ΣΕΝΑΡΙΟ_1'] = 'Α1'
                else:
                    df_result.loc[idx, 'ΤΜΗΜΑ'] = 'Α2'
                    df_step2.loc[idx, 'ΒΗΜΑ2_ΣΕΝΑΡΙΟ_1'] = 'Α2'
        else:
            st.warning("⚠️ Δεν βρέθηκε στήλη ΦΥΛΟ - απλή εναλλακτική κατανομή")
            # Απλή εναλλακτική κατανομή
            for i, idx in enumerate(remaining):
                tmima = 'Α1' if i % 2 == 0 else 'Α2'
                df_result.loc[idx, 'ΤΜΗΜΑ'] = tmima
                df_step2.loc[idx, 'ΒΗΜΑ2_ΣΕΝΑΡΙΟ_1'] = tmima
        
        # ΑΠΟΘΗΚΕΥΣΗ ΓΙΑ ΑΝΑΛΥΤΙΚΑ ΒΗΜΑΤΑ
        st.session_state.detailed_steps['ΒΗΜΑ2_ΣΕΝΑΡΙΟ_1'] = df_step2.copy()
        
        progress_bar.progress(80)
        
        # Τελικοποίηση και έλεγχος
        status_text.text("Τελικοποίηση...")
        
        # ΑΠΟΘΗΚΕΥΣΗ ΤΕΛΙΚΟΥ ΒΗΜΑΤΟΣ
        df_final = df_result.copy()
        df_final['ΒΗΜΑ6_ΣΕΝΑΡΙΟ_1'] = df_final['ΤΜΗΜΑ']
        st.session_state.detailed_steps['ΒΗΜΑ6_ΣΕΝΑΡΙΟ_1'] = df_final.copy()
        
        # Έλεγχος αποτελεσμάτων
        a1_count = len(df_result[df_result['ΤΜΗΜΑ'] == 'Α1'])
        a2_count = len(df_result[df_result['ΤΜΗΜΑ'] == 'Α2'])
        unassigned = len(df_result[df_result['ΤΜΗΜΑ'].isna()])
        
        st.write(f"**DEBUG - Τελικά αποτελέσματα:**")
        st.write(f"- Α1: {a1_count} μαθητές")
        st.write(f"- Α2: {a2_count} μαθητές") 
        st.write(f"- Ατοποθέτητοι: {unassigned} μαθητές")
        
        progress_bar.progress(100)
        status_text.text("✅ Ανάθεση ολοκληρώθηκε!")
        
        return df_result
        
    except Exception as e:
        st.error(f"Σφάλμα στην ανάθεση: {e}")
        st.code(traceback.format_exc())
        return None

def calculate_simple_score(df, tmima_col):
    """Απλός υπολογισμός score"""
    try:
        a1_data = df[df[tmima_col] == 'Α1']
        a2_data = df[df[tmima_col] == 'Α2']
        
        # Διαφορά πληθυσμού
        pop_diff = abs(len(a1_data) - len(a2_data))
        
        # Διαφορά φύλου
        a1_boys = len(a1_data[a1_data['ΦΥΛΟ'] == 'Α']) if 'ΦΥΛΟ' in df.columns else 0
        a1_girls = len(a1_data[a1_data['ΦΥΛΟ'] == 'Κ']) if 'ΦΥΛΟ' in df.columns else 0
        a2_boys = len(a2_data[a2_data['ΦΥΛΟ'] == 'Α']) if 'ΦΥΛΟ' in df.columns else 0
        a2_girls = len(a2_data[a2_data['ΦΥΛΟ'] == 'Κ']) if 'ΦΥΛΟ' in df.columns else 0
        
        boys_diff = abs(a1_boys - a2_boys)
        girls_diff = abs(a1_girls - a2_girls)
        gender_diff = max(boys_diff, girls_diff)
        
        # Διαφορά γνώσης ελληνικών
        greek_diff = 0
        if 'ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ' in df.columns:
            a1_greek = len(a1_data[a1_data['ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ'] == 'Ν'])
            a2_greek = len(a2_data[a2_data['ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ'] == 'Ν'])
            greek_diff = abs(a1_greek - a2_greek)
        
        # Συνολικό score (πολλαπλασιαστές βάσει σπουδαιότητας)
        total_score = pop_diff * 3 + gender_diff * 2 + greek_diff * 1
        
        return {
            'total_score': total_score,
            'pop_diff': pop_diff,
            'gender_diff': gender_diff,
            'greek_diff': greek_diff,
            'a1_total': len(a1_data),
            'a2_total': len(a2_data),
            'a1_boys': a1_boys,
            'a1_girls': a1_girls,
            'a2_boys': a2_boys,
            'a2_girls': a2_girls
        }
        
    except Exception as e:
        st.error(f"Σφάλμα στον υπολογισμό score: {e}")
        return None

def create_detailed_steps_workbook():
    """ΝΕΟΣ ΚΩΔΙΚΑΣ - Δημιουργία Excel workbook με όλα τα αναλυτικά βήματα"""
    try:
        excel_buffer = io.BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            # Ταξινόμηση των βημάτων για σωστή σειρά
            step_order = ['ΒΗΜΑ1', 'ΒΗΜΑ2', 'ΒΗΜΑ3', 'ΒΗΜΑ4', 'ΒΗΜΑ5', 'ΒΗΜΑ6']
            
            # Οργάνωση των sheets ανά βήμα και σενάριο
            for step in step_order:
                sheets_for_step = []
                for sheet_name, df in st.session_state.detailed_steps.items():
                    if step in sheet_name:
                        sheets_for_step.append((sheet_name, df))
                
                # Ταξινόμηση ανά σενάριο
                sheets_for_step.sort(key=lambda x: x[0])
                
                for sheet_name, df in sheets_for_step:
                    # Περιορισμός μήκους ονόματος sheet (Excel limit)
                    safe_sheet_name = sheet_name[:31] if len(sheet_name) > 31 else sheet_name
                    df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
        
        excel_buffer.seek(0)
        return excel_buffer.getvalue()
        
    except Exception as e:
        st.error(f"Σφάλμα στη δημιουργία αναλυτικών βημάτων: {e}")
        return None

def create_download_package(df, scenario_name="ΣΕΝΑΡΙΟ_1"):
    """Δημιουργία αρχείου download"""
    try:
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Κύριο αρχείο Excel
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Αποτελέσματα', index=False)
                
                # Στατιστικά αν είναι διαθέσιμα
                if STATS_AVAILABLE and 'ΤΜΗΜΑ' in df.columns:
                    try:
                        df_assigned = df[df['ΤΜΗΜΑ'].notna()].copy()
                        stats_df = generate_statistics_table(df_assigned)
                        stats_df.to_excel(writer, sheet_name='Στατιστικά', index=True)
                    except:
                        pass
            
            zip_file.writestr(f"{scenario_name}_Αποτελέσματα.xlsx", excel_buffer.getvalue())
        
        return zip_buffer.getvalue()
        
    except Exception as e:
        st.error(f"Σφάλμα στη δημιουργία αρχείου: {e}")
        return None

def main():
    """Κύρια συνάρτηση"""
    init_session_state()
    
    st.title("🎓 Σύστημα Ανάθεσης Μαθητών σε Τμήματα")
    st.markdown("*Λειτουργική έκδοση με απλή ανάθεση + Αναλυτικά Βήματα*")
    st.markdown("---")
    
    # Sidebar
    st.sidebar.title("📋 Μενού Πλοήγησης")
    
    # Debug mode toggle
    debug_mode = st.sidebar.checkbox("🔧 Debug Mode", value=True, help="Εμφάνιση debug πληροφοριών")
    
    # Upload αρχείου
    st.sidebar.subheader("📁 Φόρτωση Δεδομένων")
    uploaded_file = st.sidebar.file_uploader(
        "Επιλέξτε αρχείο Excel ή CSV",
        type=['xlsx', 'csv'],
        help="Το αρχείο πρέπει να περιέχει στήλες: ΟΝΟΜΑ, ΦΥΛΟ, ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ, ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ"
    )
    
    if uploaded_file is not None:
        # Φόρτωση δεδομένων
        if st.session_state.data is None:
            with st.spinner("Φόρτωση δεδομένων..."):
                # Προσωρινά ενεργοποιούμε το debug για τη φόρτωση
                original_debug = st.session_state.get('debug_mode', True)
                st.session_state.debug_mode = debug_mode
                
                data, error = safe_load_data(uploaded_file)
                if error:
                    st.error(f"❌ {error}")
                    return
                st.session_state.data = data
                st.session_state.current_step = 1
        
        df = st.session_state.data
        
        if df is not None:
            # Εμφάνιση βασικών στοιχείων
            if debug_mode:
                display_basic_info(df)
            else:
                # Μη-debug έκδοση
                st.subheader("📊 Βασικές Πληροφορίες")
                total_students = len(df)
                boys_count = len(df[df['ΦΥΛΟ'] == 'Α']) if 'ΦΥΛΟ' in df.columns else 0
                girls_count = len(df[df['ΦΥΛΟ'] == 'Κ']) if 'ΦΥΛΟ' in df.columns else 0
                teachers_count = len(df[df['ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ'] == 'Ν']) if 'ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ' in df.columns else 0
                
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Συνολικοί Μαθητές", total_students)
                with col2:
                    st.metric("Αγόρια", boys_count)
                with col3:
                    st.metric("Κορίτσια", girls_count)
                with col4:
                    st.metric("Παιδιά Εκπαιδευτικών", teachers_count)
            
            # Έλεγχος στηλών
            required_cols = ['ΟΝΟΜΑ', 'ΦΥΛΟ', 'ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ', 'ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ']
            missing_cols = [col for col in required_cols if col not in df.columns]
            
            if missing_cols:
                st.warning(f"⚠️ Λείπουν στήλες: {', '.join(missing_cols)}")
                if debug_mode:
                    st.info("Διαθέσιμες στήλες: " + ", ".join(df.columns))
            else:
                st.success("✅ Όλες οι απαιτούμενες στήλες βρέθηκαν!")
            
            # Store debug mode για χρήση σε άλλες συναρτήσεις
            st.session_state.debug_mode = debug_mode
            
            # Επιλογή εκτέλεσης
            st.sidebar.subheader("🚀 Εκτέλεση Ανάθεσης")
            
            if st.sidebar.button("▶️ Εκτέλεση Ανάθεσης", disabled=bool(missing_cols)):
                with st.spinner("Εκτέλεση ανάθεσης..."):
                    result_df = run_simple_assignment(df)
                    if result_df is not None:
                        st.session_state.results['final'] = result_df
                        st.session_state.current_step = 2
            
            # Εμφάνιση αποτελεσμάτων
            if 'final' in st.session_state.results:
                st.markdown("---")
                st.subheader("🏆 Αποτελέσματα Ανάθεσης")
                
                result_df = st.session_state.results['final']
                
                # Υπολογισμός και εμφάνιση score
                score = calculate_simple_score(result_df, 'ΤΜΗΜΑ')
                if score:
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("Συνολικό Score", score['total_score'])
                    with col2:
                        st.metric("Διαφορά Πληθυσμού", score['pop_diff'])
                    with col3:
                        st.metric("Διαφορά Φύλου", score['gender_diff'])
                    with col4:
                        st.metric("Διαφορά Γνώσης", score['greek_diff'])
                    
                    # Αναλυτικός πίνακας αποτελεσμάτων
                    st.subheader("📊 Αναλυτικά Αποτελέσματα")
                    summary_data = [
                        {
                            'Τμήμα': 'Α1',
                            'Συνολικός Πληθυσμός': score['a1_total'],
                            'Αγόρια': score['a1_boys'],
                            'Κορίτσια': score['a1_girls']
                        },
                        {
                            'Τμήμα': 'Α2', 
                            'Συνολικός Πληθυσμός': score['a2_total'],
                            'Αγόρια': score['a2_boys'],
                            'Κορίτσια': score['a2_girls']
                        }
                    ]
                    summary_df = pd.DataFrame(summary_data)
                    st.dataframe(summary_df, use_container_width=True)
                
                # Στατιστικά ανά τμήμα
                display_scenario_stats(result_df, 'ΤΜΗΜΑ', 'Τελικό Σενάριο')
                
                # Εμφάνιση πλήρων αποτελεσμάτων
                with st.expander("📋 Πλήρη Αποτελέσματα"):
                    st.dataframe(result_df, use_container_width=True)
                
                # Download Section
                st.sidebar.subheader("💾 Λήψη Αποτελεσμάτων")
                
                # ΚΟΥΜΠΙ 1: Δημιουργία Αρχείου (το υπάρχον)
                if st.sidebar.button("📥 Δημιουργία Αρχείου"):
                    with st.spinner("Δημιουργία αρχείου..."):
                        zip_data = create_download_package(result_df)
                        if zip_data:
                            st.sidebar.download_button(
                                label="⬇️ Λήψη Αποτελεσμάτων",
                                data=zip_data,
                                file_name="Αποτελέσματα_Ανάθεσης.zip",
                                mime="application/zip"
                            )
                
                # ΚΟΥΜΠΙ 2: Αναλυτικά Βήματα (ΤΟ ΝΕΟ ΚΟΥΜΠΙ)
                if st.sidebar.button("📋 Αναλυτικά Βήματα (VIMA6)", 
                                    help="Κατεβάστε Excel με όλα τα ενδιάμεσα βήματα"):
                    with st.spinner("Δημιουργία αρχείου αναλυτικών βημάτων..."):
                        detailed_excel = create_detailed_steps_workbook()
                        if detailed_excel:
                            st.sidebar.download_button(
                                label="⬇️ Λήψη Αναλυτικών Βημάτων",
                                data=detailed_excel,
                                file_name="VIMA6_from_ALL_SHEETS.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="detailed_steps_download"
                            )
                            st.sidebar.success("✅ Αρχείο αναλυτικών βημάτων έτοιμο!")
                        else:
                            st.sidebar.error("❌ Σφάλμα στη δημιουργία αρχείου")
            
            # Reset
            if st.sidebar.button("🔄 Επαναφορά"):
                st.session_state.clear()
                st.rerun()
    
    else:
        st.info("👆 Παρακαλώ ανεβάστε ένα αρχείο Excel ή CSV για να ξεκινήσετε")
        
        # Οδηγίες
        with st.expander("📖 Οδηγίες Χρήσης"):
            st.markdown("""
            ### Απαιτούμενες Στήλες:
            - **ΟΝΟΜΑ**: Ονοματεπώνυμο μαθητή
            - **ΦΥΛΟ**: Α (Αγόρι) ή Κ (Κορίτσι)
            - **ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ**: Ν (Ναι) ή Ο (Όχι)
            - **ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ**: Ν (Ναι) ή Ο (Όχι)
            
            ### Προαιρετικές Στήλες:
            - **ΦΙΛΟΙ**: Λίστα φίλων
            - **ΖΩΗΡΟΣ**: Ν/Ο
            - **ΙΔΙΑΙΤΕΡΟΤΗΤΑ**: Ν/Ο
            - **ΣΥΓΚΡΟΥΣΗ**: Λίστα συγκρουόμενων
            
            ### Τι κάνει η ανάθεση:
            1. Κατανέμει τα παιδιά εκπαιδευτικών ισοκατανομή
            2. Κατανέμει τους υπόλοιπους μαθητές με βάση το φύλο
            3. Υπολογίζει score βάσει ισορροπίας τμημάτων
            4. Δημιουργεί αρχεία Excel με αποτελέσματα
            
            ### 🔥 ΝΕΟ: Αναλυτικά Βήματα!
            Κατεβάστε το αρχείο **VIMA6_from_ALL_SHEETS.xlsx** που περιέχει:
            - Όλα τα ενδιάμεσα βήματα της ανάθεσης
            - Ξεχωριστό sheet για κάθε βήμα
            - Format συμβατό με το υπάρχον σύστημά σας
            """)

if __name__ == "__main__":
    main()

def create_detailed_steps_workbook():
    """Δημιουργεί ΜΟΝΟ ένα Excel με 1 φύλλο 'ΑΝΑΛΥΤΙΚΑ_ΒΗΜΑΤΑ' (VIMA6) + προαιρετικό 'VIMA7_SCORE_ΣΥΝΟΨΗ'."""
    try:
        import streamlit as st
        if getattr(st.session_state, "data", None) is None or not getattr(st.session_state, "detailed_steps", {}):
            return None
        excel_bytes = build_vima6_excel_bytes(st.session_state.data, st.session_state.detailed_steps, step7_scores=None)
        return excel_bytes
    except Exception as e:
        import streamlit as st, traceback
        st.error(f"Σφάλμα στη δημιουργία VIMA6 αρχείου: {e}")
        st.code(traceback.format_exc())
        return None



# === VIMA6 universal sidebar export button (auto-added) ===
try:
    import streamlit as st  # safe re-import
except Exception:
    pass

def _render_vima6_sidebar_button():
    try:
        st.sidebar.markdown("### 📋 Αναλυτικά Βήματα (VIMA6)")
        if getattr(st.session_state, "data", None) is None or not getattr(st.session_state, "detailed_steps", {}):
            st.sidebar.info("Φόρτωσε δεδομένα και εκτέλεσε την κατανομή για να ενεργοποιηθεί η λήψη.")
            return
        if st.sidebar.button("⬇️ VIMA6_from_ALL_SHEETS.xlsx", key="vima6_sidebar_btn", use_container_width=True):
            # Προαιρετικά: συλλογή βαθμολογιών Βήμα 7 αν υπάρχουν
            step7_scores = None
            try:
                if isinstance(st.session_state.detailed_steps, dict):
                    scores = {}
                    for scen_num, scen_data in st.session_state.detailed_steps.items():
                        if isinstance(scen_data, dict) and 'step7_score' in scen_data:
                            scores[scen_num] = scen_data['step7_score']
                    step7_scores = scores or None
            except Exception:
                step7_scores = None
            excel_bytes = build_vima6_excel_bytes(
                base_df=st.session_state.data,
                detailed_steps=st.session_state.detailed_steps,
                step7_scores=step7_scores
            )
            st.sidebar.download_button(
                label="Λήψη αρχείου",
                data=excel_bytes,
                file_name="VIMA6_from_ALL_SHEETS.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="vima6_sidebar_dl"
            )
    except Exception as _e:
        try:
            st.sidebar.error(f"Σφάλμα κουμπιού VIMA6: {_e}")
        except Exception:
            pass

# Αυτό το καλούμε ώστε να εμφανίζεται το κουμπί σε κάθε σελίδα του app
try:
    _render_vima6_sidebar_button()
except Exception:
    pass
