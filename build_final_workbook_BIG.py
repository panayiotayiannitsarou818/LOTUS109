
# -*- coding: utf-8 -*-
"""
build_final_workbook_BIG.py
- Ανθεκτικό σε μεγάλα workbooks
- Παράγει: Sx_SCORE, SUMMARY (με ένδειξη νικητή), FINAL_SCENARIO, FINAL_SCENARIO_AtoP
"""
from __future__ import annotations
import argparse, re, random
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import step7

RANDOM_SEED = 42
random.seed(RANDOM_SEED)

BASE_COLS = ["ΟΝΟΜΑ", "ΦΥΛΟ", "ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ", "ΖΩΗΡΟΣ", "ΙΔΙΑΙΤΕΡΟΤΗΤΑ", "ΦΙΛΟΙ"]

def sheet_names(input_path, pattern=r"^S\\d+$"):
    xls = pd.ExcelFile(input_path)
    rgx = re.compile(pattern) if pattern else None
    return [s for s in xls.sheet_names if (rgx.match(str(s)) if rgx else True)]

def detect_scenario_col(header_df: pd.DataFrame, sid: int) -> str:
    cand = f"ΒΗΜΑ6_ΣΕΝΑΡΙΟ_{sid}"
    if cand in header_df.columns: return cand
    cands = [c for c in header_df.columns if re.match(r"^ΒΗΜΑ6_ΣΕΝΑΡΙΟ_\\d+", str(c))]
    if cands: return cands[0]
    for alt in ("ΒΗΜΑ6_ΤΜΗΜΑ", "ΤΜΗΜΑ_ΜΕΤΑ_ΒΗΜΑ6", "ΤΜΗΜΑ"):
        if alt in header_df.columns: return alt
    raise ValueError("Δεν βρέθηκε στήλη Βήματος 6 στο φύλλο.")

def copy_values_only(in_path, out_path):
    wb_in = load_workbook(in_path, read_only=True, data_only=True)
    wb_out = Workbook()
    # remove default
    if "Sheet" in wb_out.sheetnames:
        std = wb_out["Sheet"]; wb_out.remove(std)
    for ws_in in wb_in.worksheets:
        ws_out = wb_out.create_sheet(ws_in.title)
        for row in ws_in.iter_rows(values_only=True):
            ws_out.append(list(row))
    wb_out.save(out_path)

def process(input_path, output_path, pattern=r"^S\\d+$"):
    copy_values_only(input_path, output_path)

    writer = pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace")
    summary_rows = []

    for sheet in sheet_names(input_path, pattern):
        header_df = pd.read_excel(input_path, sheet_name=sheet, nrows=0)
        sid_m = re.search(r"(\\d+)", str(sheet)); sid = int(sid_m.group(1)) if sid_m else 1
        scen_col = detect_scenario_col(header_df, sid)

        needed = list(dict.fromkeys([*BASE_COLS, scen_col]))
        present = [c for c in needed if c in header_df.columns]
        df = pd.read_excel(input_path, sheet_name=sheet, usecols=present)

        for opt in ("ΖΩΗΡΟΣ", "ΙΔΙΑΙΤΕΡΟΤΗΤΑ", "ΦΙΛΟΙ"):
            if opt not in df.columns: df[opt] = ""

        df[scen_col] = df[scen_col].astype(str).str.strip().str.replace(r"^A", "Α", regex=True)

        s = step7.score_one_scenario(df, scen_col)

        row = {
            "SCENARIO": sheet,
            "SCENARIO_COL": scen_col,
            "TOTAL": s["total_score"],
            "POP_DIFF": s["diff_population"],
            "BOYS_DIFF": s["diff_boys"],
            "GIRLS_DIFF": s["diff_girls"],
            "GENDER_DIFF_TOTAL": s["diff_boys"] + s["diff_girls"],
            "GREEK_DIFF": s["diff_greek"],
            "POP_PEN": s["population_penalty"],
            "BOYS_PEN": s["boys_penalty"],
            "GIRLS_PEN": s["girls_penalty"],
            "GENDER_PEN": s["gender_penalty"],
            "GREEK_PEN": s["greek_penalty"],
            "CONFLICT_PEN": s["conflict_penalty"],
            "BROKEN_COUNT": s["broken_friendships"],
            "BROKEN_PEN": s["broken_friendships_penalty"],
        }
        summary_rows.append(row)

        pd.DataFrame([row]).to_excel(writer, index=False, sheet_name=f"{sheet}_SCORE", startrow=0)

    summary_df = pd.DataFrame(summary_rows).sort_values(
        by=["TOTAL", "POP_DIFF", "GENDER_DIFF_TOTAL", "GREEK_DIFF"],
        ascending=[True, True, True, True]
    ).reset_index(drop=True)

    top_mask = (summary_df["TOTAL"] == summary_df.loc[0,"TOTAL"]) & \
               (summary_df["POP_DIFF"] == summary_df.loc[0,"POP_DIFF"]) & \
               (summary_df["GENDER_DIFF_TOTAL"] == summary_df.loc[0,"GENDER_DIFF_TOTAL"]) & \
               (summary_df["GREEK_DIFF"] == summary_df.loc[0,"GREEK_DIFF"])
    top_group = summary_df[top_mask].copy()
    if len(top_group) > 1:
        random.seed(RANDOM_SEED)
        best_idx = random.choice(list(top_group.index))
    else:
        best_idx = 0
    summary_df["IS_BEST"] = False
    summary_df.loc[best_idx, "IS_BEST"] = True
    summary_df.to_excel(writer, index=False, sheet_name="SUMMARY")
    writer.close()

    wb = load_workbook(output_path)
    ws = wb["SUMMARY"]

    headers = [cell.value for cell in ws[1]]
    scen_idx = headers.index("SCENARIO")+1
    isbest_idx = headers.index("IS_BEST")+1

    winner_row = None
    for r in range(2, ws.max_row+1):
        if str(ws.cell(row=r, column=isbest_idx).value) == "True":
            winner_row = r; break
    winner_name = ws.cell(row=winner_row, column=scen_idx).value if winner_row else summary_df.loc[0,"SCENARIO"]

    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    font_bold = Font(bold=True)
    thin = Side(border_style="thin", color="9BBB59")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for c in range(1, ws.max_column+1):
        cell = ws.cell(row=winner_row, column=c)
        cell.fill = fill; cell.font = font_bold; cell.border = border; cell.alignment = Alignment(vertical="center")
    note_col = ws.max_column + 2
    ws.cell(row=1, column=note_col, value="FINAL WINNER").font = Font(bold=True)
    ws.cell(row=winner_row, column=note_col, value="✅").font = Font(bold=True)
    ws.freeze_panes = "A2"

    # FULL copy
    if "FINAL_SCENARIO" in wb.sheetnames:
        del wb["FINAL_SCENARIO"]
    ws_src = wb[winner_name]
    ws_full = wb.create_sheet("FINAL_SCENARIO")
    for row in ws_src.iter_rows(values_only=True):
        ws_full.append(list(row))
    for cell in ws_full[1]:
        cell.font = Font(bold=True); cell.alignment = Alignment(vertical="center")

    # A..P copy (first 16 columns)
    if "FINAL_SCENARIO_AtoP" in wb.sheetnames:
        del wb["FINAL_SCENARIO_AtoP"]
    ws_ap = wb.create_sheet("FINAL_SCENARIO_AtoP")
    max_rows = ws_src.max_row
    max_cols = min(16, ws_src.max_column)
    ws_ap.append([ws_src.cell(row=1, column=c).value for c in range(1, max_cols+1)])
    for cell in ws_ap[1]:
        cell.font = Font(bold=True); cell.alignment = Alignment(vertical="center")
    for r in range(2, max_rows+1):
        ws_ap.append([ws_src.cell(row=r, column=c).value for c in range(1, max_cols+1)])
    ws_ap.freeze_panes = "A2"
    for col_idx in range(1, max_cols+1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for r in range(1, ws_ap.max_row+1):
            v = ws_ap.cell(row=r, column=col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws_ap.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(output_path)
    return output_path

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Excel από Βήμα 6 (πολλά φύλλα Sx)")
    ap.add_argument("--output", required=True, help="Τελικό Excel")
    ap.add_argument("--sheet_regex", default=r"^S\\d+$", help="Regex για φύλλα-σενάρια (default S1,S2,...)")
    args = ap.parse_args()
    out = process(args.input, args.output, pattern=args.sheet_regex)
    print(out)

if __name__ == "__main__":
    main()
